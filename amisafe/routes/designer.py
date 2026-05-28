"""웹 기반 양식 디자이너 (form_designer_pro_ui_excel.py의 웹 포팅).

데스크톱 Tkinter 앱을 HTML5 Canvas + JS로 재구현한다.
관리자만 접근 가능하다.

엔드포인트:
- GET  /designer                                 디자이너 UI
- GET  /designer/api/config                      현재 form_config.json
- POST /designer/api/config                      전체 form_config 저장
- POST /designer/api/import-category-excel       대/소분류 엑셀에서 dropdown options 추출
- GET  /designer/api/excel-template              엑셀 템플릿 다운로드
- POST /designer/api/auto-detect-fields          OpenAI Vision으로 입력 영역 자동 인식
"""
import base64
import json
import os
from io import BytesIO
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from flask import (
    Blueprint, render_template, request, redirect, url_for, session,
    jsonify, send_file,
)

from amisafe.config import Config
from amisafe.utils import allowed_image_filename
from amisafe.services.form_config import load_form_config, save_form_config

bp = Blueprint("designer", __name__, url_prefix="/designer")


def _require_admin_json():
    user = session.get("user")
    if not user:
        return None, (jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401)
    if not user.get("is_admin"):
        return None, (jsonify({"ok": False, "message": "관리자만 접근 가능합니다."}), 403)
    return user, None


@bp.route("")
def designer_page():
    """디자이너 메인 페이지."""
    user = session.get("user")
    if not user:
        return redirect(url_for("auth.home"))
    if not user.get("is_admin"):
        return "관리자만 접근 가능합니다.", 403

    image_files = sorted(os.listdir(Config.FORMS_FOLDER)) if os.path.exists(Config.FORMS_FOLDER) else []
    image_files = [f for f in image_files if allowed_image_filename(f)]

    return render_template("designer.html", user=user, image_files=image_files)


@bp.route("/api/config", methods=["GET"])
def api_get_config():
    """현재 form_config.json 반환."""
    user, err = _require_admin_json()
    if err is not None:
        return err
    return jsonify({"ok": True, "config": load_form_config()})


@bp.route("/api/config", methods=["POST"])
def api_save_config():
    """form_config.json 전체 저장 (구조 검증 후)."""
    user, err = _require_admin_json()
    if err is not None:
        return err

    payload = request.get_json(silent=True) or {}
    if "forms" not in payload or not isinstance(payload["forms"], list):
        return jsonify({"ok": False, "message": "forms 배열이 필요합니다."}), 400

    try:
        save_form_config(payload)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "message": f"저장 실패: {e}"}), 500


@bp.route("/api/upload-image", methods=["POST"])
def api_upload_image():
    """디자이너에서 배경 이미지 업로드."""
    user, err = _require_admin_json()
    if err is not None:
        return err

    file = request.files.get("image_file")
    if not file or not file.filename:
        return jsonify({"ok": False, "message": "파일이 없습니다."}), 400

    original_name = os.path.basename(file.filename)
    if not allowed_image_filename(original_name):
        return jsonify({"ok": False, "message": "이미지 파일만 업로드 가능합니다."}), 400

    save_path = os.path.join(Config.FORMS_FOLDER, original_name)
    file.save(save_path)
    return jsonify({"ok": True, "filename": original_name})


@bp.route("/api/import-category-excel", methods=["POST"])
def api_import_category_excel():
    """엑셀(대분류/소분류)에서 dropdown options 추출.

    원본 form_designer.py의 _import_category_excel / _read_category_excel 와 동일 동작.
    A열=대분류, B열=소분류 형태를 읽어 parent_options[]와 option_map{대분류: [소분류]}을 반환.
    """
    user, err = _require_admin_json()
    if err is not None:
        return err

    file = request.files.get("excel_file")
    if not file or not file.filename:
        return jsonify({"ok": False, "message": "엑셀 파일이 없습니다."}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"ok": False, "message": "openpyxl이 설치되지 않았습니다."}), 500

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active

        parent_options = []
        seen_parents = set()
        option_map = {}
        rows_read = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(c is None for c in row):
                continue
            parent = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            child = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not parent:
                continue

            rows_read += 1
            if parent not in seen_parents:
                seen_parents.add(parent)
                parent_options.append({"option_label": parent, "option_value": parent})
                option_map[parent] = []

            if child and child not in option_map[parent]:
                option_map[parent].append(child)

        # 소분류는 option_label/option_value 두 키로 변환
        normalized_map = {
            parent: [{"option_label": c, "option_value": c} for c in children]
            for parent, children in option_map.items()
        }

        return jsonify({
            "ok": True,
            "parent_options": parent_options,
            "option_map": normalized_map,
            "stats": {
                "rows_read": rows_read,
                "parent_count": len(parent_options),
                "child_total": sum(len(v) for v in normalized_map.values()),
            },
        })
    except Exception as e:
        return jsonify({"ok": False, "message": f"엑셀 읽기 실패: {e}"}), 400


@bp.route("/api/excel-template", methods=["GET"])
def api_excel_template():
    """대/소분류 엑셀 템플릿 다운로드."""
    user = session.get("user")
    if not user or not user.get("is_admin"):
        return "관리자만 접근 가능합니다.", 403

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl이 설치되지 않았습니다.", 500

    wb = Workbook()
    ws = wb.active
    ws.title = "categories"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "대분류"
    ws["B1"] = "소분류"
    for cell in (ws["A1"], ws["B1"]):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ("점검항목", "정상"), ("점검항목", "이상"), ("점검항목", "해당없음"),
        ("작업유형", "고소작업"), ("작업유형", "용접작업"), ("작업유형", "전기작업"),
    ]
    for row_idx, (a, b) in enumerate(samples, start=2):
        ws.cell(row=row_idx, column=1, value=a)
        ws.cell(row=row_idx, column=2, value=b)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="대소분류_목록박스_양식.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


_VISION_PROMPT = (
    "이 이미지는 한국 산업 안전점검 양식입니다. 사용자가 손으로 채워 넣어야 하는 "
    "'입력 영역'만 찾아 위치를 알려주세요. 인쇄된 고정 라벨, 제목, 설명 텍스트는 "
    "포함하지 마세요.\n\n"
    "각 입력 영역에 대해 다음을 출력합니다:\n"
    "- type: 'text'(자유 입력), 'date'(날짜), 'checkbox'(체크박스), 'signature'(서명) 중 하나\n"
    "- label: 그 영역 옆/위/제목 칼럼에 적힌 한국어 라벨 (예: '작업자명', '일자', '확인')\n"
    "- x_pct, y_pct: 영역의 좌상단 좌표를 이미지 너비/높이 대비 0.0~1.0 비율로\n"
    "- w_pct, h_pct: 영역의 너비/높이를 이미지 너비/높이 대비 0.0~1.0 비율로\n\n"
    "체크박스는 정사각형 작은 영역(보통 w_pct, h_pct가 비슷한 작은 값). "
    "서명란은 가로로 긴 빈 영역(보통 '서명', '확인' 라벨 옆).\n"
    "반드시 JSON 객체 하나만 출력하세요. 형식:\n"
    '{"fields": [{"type": "...", "label": "...", "x_pct": 0.12, "y_pct": 0.30, "w_pct": 0.20, "h_pct": 0.04}, ...]}'
)

_ALLOWED_TYPES = {"text", "date", "checkbox", "signature"}
_MIME_BY_EXT = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png", ".webp": "image/webp",
}


def _sanitize_detected_field(raw):
    """모델 응답의 한 필드를 안전한 dict로 변환. 잘못된 항목은 None."""
    try:
        ftype = str(raw.get("type", "")).strip().lower()
        if ftype not in _ALLOWED_TYPES:
            return None
        return {
            "type": ftype,
            "label": str(raw.get("label", "") or "").strip()[:50],
            "x_pct": max(0.0, min(1.0, float(raw.get("x_pct", 0)))),
            "y_pct": max(0.0, min(1.0, float(raw.get("y_pct", 0)))),
            "w_pct": max(0.005, min(1.0, float(raw.get("w_pct", 0.05)))),
            "h_pct": max(0.005, min(1.0, float(raw.get("h_pct", 0.04)))),
        }
    except (TypeError, ValueError):
        return None


@bp.route("/api/auto-detect-fields", methods=["POST"])
def api_auto_detect_fields():
    """업로드된 배경 이미지를 OpenAI Vision에 보내 입력 영역을 자동 인식한다.

    요청: {filename: "<forms/ 하위 이미지 파일명>"}
    응답: {ok: true, fields: [{type, label, x_pct, y_pct, w_pct, h_pct}, ...]}
    """
    user, err = _require_admin_json()
    if err is not None:
        return err

    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return jsonify({
            "ok": False,
            "message": "서버에 OPENAI_API_KEY가 설정돼 있지 않습니다. (Render 환경변수)",
        }), 500

    payload = request.get_json(silent=True) or {}
    filename = os.path.basename(str(payload.get("filename", "")).strip())
    if not filename:
        return jsonify({"ok": False, "message": "filename이 필요합니다."}), 400

    image_path = os.path.join(Config.FORMS_FOLDER, filename)
    if not os.path.isfile(image_path):
        return jsonify({"ok": False, "message": "이미지 파일을 찾을 수 없습니다."}), 404

    ext = os.path.splitext(filename)[1].lower()
    mime = _MIME_BY_EXT.get(ext, "image/jpeg")

    with open(image_path, "rb") as fp:
        b64 = base64.b64encode(fp.read()).decode("ascii")
    data_url = f"data:{mime};base64,{b64}"

    api_payload = {
        "model": os.environ.get("OPENAI_VISION_MODEL", "gpt-4o-mini"),
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": _VISION_PROMPT},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0.1,
        "max_tokens": 2000,
    }

    req = Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(api_payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )

    try:
        with urlopen(req, timeout=90) as resp:
            api_raw = resp.read().decode("utf-8")
        api_data = json.loads(api_raw)
    except HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")[:300]
        except Exception:
            pass
        return jsonify({"ok": False, "message": f"OpenAI 호출 실패 ({e.code}): {body}"}), 502
    except URLError as e:
        return jsonify({"ok": False, "message": f"네트워크 오류: {e.reason}"}), 502
    except Exception as e:
        return jsonify({"ok": False, "message": f"오류: {e}"}), 500

    try:
        content = api_data["choices"][0]["message"]["content"]
        parsed = json.loads(content)
    except (KeyError, IndexError, TypeError, ValueError) as e:
        return jsonify({"ok": False, "message": f"AI 응답 파싱 실패: {e}"}), 502

    raw_fields = parsed.get("fields", []) if isinstance(parsed, dict) else []
    detected = []
    for item in raw_fields:
        if isinstance(item, dict):
            cleaned = _sanitize_detected_field(item)
            if cleaned:
                detected.append(cleaned)

    return jsonify({"ok": True, "fields": detected, "count": len(detected)})
